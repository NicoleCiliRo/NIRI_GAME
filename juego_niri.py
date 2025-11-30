"""
JUEGO DE DETECCIÓN DE CARIES EN IMÁGENES NIRI - VERSIÓN FINAL
Juego educativo para entrenar la detección de caries en imágenes de infrarrojo cercano
Autor: Sistema de entrenamiento odontológico - Nicole Rodrigues

ARCHIVO: juego_caries.py
"""

# ============================================================================
# IMPORTACIÓN DE LIBRERÍAS
# ============================================================================

import pygame
import pygame_menu
from pygame_menu import sound
import json
import sys
import math
import random
from datetime import datetime
import os
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# ============================================================================
# INICIALIZACIÓN DE PYGAME
# ============================================================================

pygame.init()
pygame.font.init()
pygame.mixer.init()

# ============================================================================
# CONSTANTES DEL JUEGO
# ============================================================================

ANCHO_VENTANA = 1380
ALTO_VENTANA = 820

COLOR_BLANCO = (255, 255, 255)
COLOR_NEGRO = (0, 0, 0)
COLOR_AZUL = (59, 130, 246)
COLOR_AZUL_OSCURO = (30, 58, 138)
COLOR_AZUL_CLARO = (96, 165, 250)
COLOR_VERDE = (34, 197, 94)
COLOR_VERDE_HOVER = (60, 219, 124)
COLOR_ROJO = (239, 68, 68)
COLOR_AMARILLO = (250, 204, 21)
COLOR_GRIS = (148, 163, 184)
COLOR_GRIS_OSCURO = (71, 85, 105)
COLOR_FONDO = (15, 23, 42)
COLOR_PANEL = (30, 41, 59)
COLOR_PANEL_HOVER = (51, 65, 85)
COLOR_TEXTO_SECUNDARIO = (203, 213, 225)

ESTADO_MENU = "menu"
ESTADO_JUGANDO = "jugando"
ESTADO_RESULTADOS = "resultados"

EXPERIENCIA_PRINCIPIANTE = "principiante"
EXPERIENCIA_AVANZADO = "avanzado"

DIFICULTAD_FACIL = "easy"
DIFICULTAD_MEDIA = "medium"
DIFICULTAD_DIFICIL = "hard"

# ============================================================================
# CLASE PRINCIPAL DEL JUEGO
# ============================================================================

class JuegoDeteccionCaries:
    """Clase principal que controla todo el juego"""
    
    def __init__(self):
        """Constructor: inicializa todas las variables"""
        
        self.ventana = pygame.display.set_mode((ANCHO_VENTANA, ALTO_VENTANA))
        pygame.display.set_caption("🦷 Juego de Detección de Caries en imágenes NIRI")
        
        self.reloj = pygame.time.Clock()
        self.estado = ESTADO_MENU
        
        self.nombre_jugador = ""
        self.experiencia = None
        
        self.puntos = 0
        self.vidas = 5
        self.racha = 0
        self.racha_maxima = 0
        self.tiempo_inicio = 0
        self.tiempo_actual = 0
        
        self.pregunta_actual = 0
        self.respondida = False
        self.mostrar_feedback = False
        
        self.puntos_poligono = []
        self.datos_juego = []
        self.imagenes_cargadas = {}
        self.resultados_detallados = []
        
        self.mensaje_feedback = ""
        self.es_correcto = False
        self.precision_actual = 0.0
        self.fuente_titulo = pygame.font.Font(None, 60)
        self.fuente_grande = pygame.font.Font(None, 52)
        self.fuente_mediana = pygame.font.Font(None, 36)
        self.fuente_pequena = pygame.font.Font(None, 26)
        
        self.feedback_alpha = 0
        self.feedback_timer = 0
        
        self.ranking = []
        self.cargar_ranking()
        
        self.input_activo = False
        
        self.fondo_imagen = None
        self.cargar_fondo()
        
        self.tooth_icon = None
        self.cargar_icono_diente()
        
        self.musica_cargada = False
        self.cargar_musica()
        
        self.archivo_excel = "datos_juego_caries.xlsx"
        self.inicializar_excel()
        
        self.cargar_datos_desde_json()
    
    def dibujar_panel_moderno(self, rect, color_fondo, radio=20, sombra=True, intensidad_sombra=40):
        """Dibuja un panel con esquinas redondeadas y sombra suave"""
        if sombra:
            # Sombra suave con múltiples capas
            for i in range(3):
                offset = 2 + i * 2
                alpha = intensidad_sombra - i * 10
                sombra_rect = rect.copy()
                sombra_rect.x += offset
                sombra_rect.y += offset
                sombra_surf = pygame.Surface((sombra_rect.width, sombra_rect.height), pygame.SRCALPHA)
                pygame.draw.rect(sombra_surf, (0, 0, 0, alpha), sombra_surf.get_rect(), border_radius=radio)
                self.ventana.blit(sombra_surf, (sombra_rect.x, sombra_rect.y))
        
        # Panel principal con bordes redondeados
        pygame.draw.rect(self.ventana, color_fondo, rect, border_radius=radio)
    
    def dibujar_panel_glass(self, rect, alpha=180):
        """Panel estilo glassmorphism minimalista"""
        superficie = pygame.Surface((rect.width, rect.height), pygame.SRCALPHA)
        # Fondo semi-transparente más sutil
        superficie.fill((255, 255, 255, alpha // 8))
        
        # Borde sutil
        pygame.draw.rect(superficie, (255, 255, 255, alpha // 4), 
                         superficie.get_rect(), width=1, border_radius=20)
        
        self.ventana.blit(superficie, (rect.x, rect.y))
    
    def dibujar_gradiente(self, superficie, rect, color_inicio, color_fin, horizontal=False):
        """Dibuja un gradiente suave"""
        if horizontal:
            for i in range(rect.width):
                ratio = i / rect.width
                color = tuple(int(color_inicio[j] + (color_fin[j] - color_inicio[j]) * ratio) for j in range(3))
                pygame.draw.line(superficie, color, (rect.x + i, rect.y), (rect.x + i, rect.y + rect.height))
        else:
            for i in range(rect.height):
                ratio = i / rect.height
                color = tuple(int(color_inicio[j] + (color_fin[j] - color_inicio[j]) * ratio) for j in range(3))
                pygame.draw.line(superficie, color, (rect.x, rect.y + i), (rect.x + rect.width, rect.y + i))
    
    def dibujar_diente(self, x, y, size, color):
        """Dibuja un icono de diente (usa imagen o fallback)"""
        if self.tooth_icon:
            # Escalar el icono al tamaño deseado
            scaled_icon = pygame.transform.scale(self.tooth_icon, (size, size))
            self.ventana.blit(scaled_icon, (x, y))
        else:
            # Fallback: dibujar diente con formas
            corona_points = [
                (x + size * 0.2, y + size * 0.4),
                (x + size * 0.3, y + size * 0.1),
                (x + size * 0.4, y + size * 0.3),
                (x + size * 0.5, y),
                (x + size * 0.6, y + size * 0.3),
                (x + size * 0.7, y + size * 0.1),
                (x + size * 0.8, y + size * 0.4),
            ]
            pygame.draw.polygon(self.ventana, color, corona_points)
            raiz_rect = pygame.Rect(x + size * 0.25, y + size * 0.35, size * 0.5, size * 0.55)
            pygame.draw.ellipse(self.ventana, color, raiz_rect)
    
    def dibujar_corazon(self, x, y, size, color, filled=True):
        """Dibuja un corazón bonito y suave"""
        # Crear superficie para el corazón
        heart_surface = pygame.Surface((size, size), pygame.SRCALPHA)
        
        # Dibujar corazón usando curvas Bézier aproximadas con círculos
        center_x = size // 2
        
        # Dos círculos superiores (lóbulos del corazón)
        left_circle = (center_x - size // 4, size // 3)
        right_circle = (center_x + size // 4, size // 3)
        radius = size // 3
        
        pygame.draw.circle(heart_surface, color, left_circle, radius)
        pygame.draw.circle(heart_surface, color, right_circle, radius)
        
        # Triángulo inferior (punta del corazón)
        points = [
            (center_x - size // 2 + 2, size // 3),
            (center_x, size - 2),
            (center_x + size // 2 - 2, size // 3)
        ]
        pygame.draw.polygon(heart_surface, color, points)
        
        # Rectángulo central para rellenar huecos
        pygame.draw.rect(heart_surface, color, (center_x - size // 3, size // 3, size // 2, size // 3))
        
        # Blit en la ventana
        self.ventana.blit(heart_surface, (x, y))
    
    def cargar_fondo(self):
        """Carga la imagen de fondo de la pantalla"""
        try:
            if os.path.exists('fondo_pantalla.jpg'):
                self.fondo_imagen = pygame.image.load('fondo_pantalla.jpg')
                self.fondo_imagen = pygame.transform.scale(self.fondo_imagen, (ANCHO_VENTANA, ALTO_VENTANA))
                print("✅ Imagen de fondo cargada")
            else:
                print("⚠️  No se encontró fondo_pantalla.jpg (opcional)")
                self.fondo_imagen = None
        except Exception as e:
            print(f"⚠️  Error al cargar fondo: {e}")
            self.fondo_imagen = None
    
    def cargar_icono_diente(self):
        """Carga el icono de diente"""
        try:
            if os.path.exists('tooth.png'):
                self.tooth_icon = pygame.image.load('tooth.png')
                print("✅ Icono de diente cargado")
            else:
                print("⚠️  No se encontró tooth.png")
                self.tooth_icon = None
        except Exception as e:
            print(f"⚠️  Error al cargar icono: {e}")
            self.tooth_icon = None
    
    def cargar_musica(self):
        """Carga y reproduce la música de fondo en bucle"""
        try:
            if os.path.exists('soundtrak_caries.mp3'):
                pygame.mixer.music.load('soundtrak_caries.mp3')
                pygame.mixer.music.set_volume(0.3)
                pygame.mixer.music.play(-1)
                self.musica_cargada = True
                print("✅ Música de fondo cargada y reproduciendo")
            else:
                print("⚠️  No se encontró soundtrak_caries.mp3 (opcional)")
        except Exception as e:
            print(f"⚠️  Error al cargar música: {e}")
    
    def inicializar_excel(self):
        """Crea o carga el archivo Excel para guardar datos"""
        try:
            if os.path.exists(self.archivo_excel):
                self.workbook = openpyxl.load_workbook(self.archivo_excel)
                print(f"✅ Excel cargado: {self.archivo_excel}")
            else:
                self.workbook = Workbook()
                self.crear_hojas_excel()
                self.workbook.save(self.archivo_excel)
                print(f"✅ Excel creado: {self.archivo_excel}")
        except Exception as e:
            print(f"⚠️  Error con Excel: {e}")
            self.workbook = Workbook()
    
    def crear_hojas_excel(self):
        """Crea las hojas del Excel con formato"""
        if 'Sheet' in self.workbook.sheetnames:
            del self.workbook['Sheet']
        
        ws_partidas = self.workbook.create_sheet('Partidas', 0)
        headers_partidas = ['Fecha', 'Hora', 'Jugador', 'Experiencia', 'Puntos', 'Precisión %', 'Vidas Restantes', 'Racha Máxima', 'Tiempo Total (seg)', 'Preguntas Totales', 'Aciertos']
        ws_partidas.append(headers_partidas)
        
        ws_respuestas = self.workbook.create_sheet('Respuestas', 1)
        headers_respuestas = ['Fecha', 'Jugador', 'Pregunta #', 'Imagen', 'Dificultad', 'Correcto', 'Precisión %', 'Puntos', 'Tiempo (seg)']
        ws_respuestas.append(headers_respuestas)
        
        ws_ranking = self.workbook.create_sheet('Ranking', 2)
        headers_ranking = ['Posición', 'Jugador', 'Puntos', 'Precisión %', 'Experiencia', 'Fecha']
        ws_ranking.append(headers_ranking)
        
        for ws in [ws_partidas, ws_respuestas, ws_ranking]:
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def guardar_partida_excel(self):
        """Guarda los datos de la partida en Excel"""
        try:
            total_preguntas = len(self.resultados_detallados)
            aciertos = sum(1 for r in self.resultados_detallados if r['correcto'])
            precision = (aciertos / total_preguntas * 100) if total_preguntas > 0 else 0
            
            fecha = datetime.now().strftime("%Y-%m-%d")
            hora = datetime.now().strftime("%H:%M:%S")
            
            ws_partidas = self.workbook['Partidas']
            ws_partidas.append([fecha, hora, self.nombre_jugador, self.experiencia, self.puntos, round(precision, 1), self.vidas, self.racha_maxima, int(self.tiempo_actual), total_preguntas, aciertos])
            
            ws_respuestas = self.workbook['Respuestas']
            tiempo_por_pregunta = self.tiempo_actual / total_preguntas if total_preguntas > 0 else 0
            
            for resultado in self.resultados_detallados:
                nombre_imagen = self.datos_juego[resultado['pregunta'] - 1]['imageName']
                dificultad = self.datos_juego[resultado['pregunta'] - 1]['difficulty']
                ws_respuestas.append([fecha, self.nombre_jugador, resultado['pregunta'], nombre_imagen, dificultad, 'SÍ' if resultado['correcto'] else 'NO', resultado['precision'], resultado['puntos'], round(tiempo_por_pregunta, 1)])
            
            self.actualizar_ranking_excel()
            self.workbook.save(self.archivo_excel)
            print(f"✅ Datos guardados en Excel: {self.archivo_excel}")
            
        except Exception as e:
            print(f"⚠️  Error al guardar en Excel: {e}")
    
    def actualizar_ranking_excel(self):
        """Actualiza la hoja de ranking en Excel"""
        try:
            ws_ranking = self.workbook['Ranking']
            ws_ranking.delete_rows(2, ws_ranking.max_row)
            
            for idx, entrada in enumerate(self.ranking, start=1):
                ws_ranking.append([idx, entrada['nombre'], entrada['puntos'], entrada['precision'], entrada['experiencia'], entrada['fecha']])
        except Exception as e:
            print(f"⚠️  Error al actualizar ranking en Excel: {e}")
    
    def cargar_datos_desde_json(self):
        """Carga las imágenes etiquetadas desde el archivo JSON"""
        print("📦 Cargando imágenes desde JSON...")
        
        archivo_json = 'etiquetas_caries.json'
        
        if not os.path.exists(archivo_json):
            print(f"❌ No se encontró: {archivo_json}")
            print("📁 Usando imágenes de respaldo...")
            self.cargar_datos_simulados()
            return
        
        try:
            with open(archivo_json, 'r', encoding='utf-8') as f:
                datos = json.load(f)
            
            print(f"✅ Archivo JSON cargado: {len(datos)} imágenes")
            
            for dato in datos:
                nombre_imagen = dato['imageName']
                ruta_imagen = os.path.join('imagenes', nombre_imagen)
                
                if not os.path.exists(ruta_imagen):
                    print(f"⚠️  No se encontró: {ruta_imagen}")
                    continue
                
                try:
                    imagen = pygame.image.load(ruta_imagen)
                    self.imagenes_cargadas[dato['imageName']] = imagen
                    self.datos_juego.append(dato)
                    print(f"✅ Cargada: {nombre_imagen} ({dato['difficulty']})")
                except Exception as e:
                    print(f"❌ Error cargando {nombre_imagen}: {e}")
            
            print(f"\n🎉 Total cargadas: {len(self.datos_juego)} imágenes")
            
            if len(self.datos_juego) == 0:
                print("⚠️  No se cargaron imágenes, usando respaldo")
                self.cargar_datos_simulados()
        
        except Exception as e:
            print(f"❌ Error al leer JSON: {e}")
            self.cargar_datos_simulados()
    
    def cargar_datos_simulados(self):
        """Método de respaldo: crea imágenes simuladas"""
        datos_ejemplo = []
        
        configuraciones = [
            (600, 400, (300, 130), 70, DIFICULTAD_FACIL, "Simulada 1 - Caries Superior"),
            (600, 400, (175, 195), 65, DIFICULTAD_FACIL, "Simulada 2 - Caries Izquierda"),
            (600, 400, (210, 290), 60, DIFICULTAD_MEDIA, "Simulada 3 - Caries Centro-Inferior"),
        ]
        
        for idx, (ancho, alto, pos_caries, tam_caries, dificultad, nombre) in enumerate(configuraciones):
            superficie = pygame.Surface((ancho, alto))
            superficie.fill((180, 180, 180))
            
            for _ in range(500):
                x = random.randint(0, ancho)
                y = random.randint(0, alto)
                color_punto = random.randint(160, 200)
                pygame.draw.circle(superficie, (color_punto, color_punto, color_punto), (x, y), 1)
            
            centro_x, centro_y = pos_caries
            color_caries = (60, 60, 60)
            pygame.draw.circle(superficie, color_caries, (centro_x, centro_y), tam_caries)
            
            for r in range(tam_caries, 0, -5):
                intensidad = 60 + int((tam_caries - r) * 0.8)
                intensidad = min(intensidad, 100)
                pygame.draw.circle(superficie, (intensidad, intensidad, intensidad), (centro_x, centro_y), r, 2)
            
            self.imagenes_cargadas[nombre] = superficie
            
            num_puntos = 12
            poligono_correcto = []
            for i in range(num_puntos):
                angulo = (2 * math.pi * i) / num_puntos
                radio_poligono = tam_caries * 0.85
                x = centro_x + radio_poligono * math.cos(angulo)
                y = centro_y + radio_poligono * math.sin(angulo)
                poligono_correcto.append({'x': x, 'y': y})
            
            dato = {'imageName': nombre, 'difficulty': dificultad, 'polygons': [poligono_correcto], 'timestamp': datetime.now().strftime("%Y-%m-%d %H:%M:%S")}
            datos_ejemplo.append(dato)
        
        self.datos_juego = datos_ejemplo
        print(f"✅ {len(self.datos_juego)} imágenes simuladas cargadas")
    
    def cargar_ranking(self):
        """Carga el ranking desde archivo JSON"""
        try:
            with open('ranking.json', 'r', encoding='utf-8') as archivo:
                self.ranking = json.load(archivo)
        except FileNotFoundError:
            self.ranking = []
        except Exception as e:
            print(f"Error al cargar ranking: {e}")
            self.ranking = []
    
    def guardar_ranking(self):
        """Guarda el ranking en archivo JSON"""
        try:
            with open('ranking.json', 'w', encoding='utf-8') as archivo:
                json.dump(self.ranking, archivo, indent=2, ensure_ascii=False)
        except Exception as e:
            print(f"Error al guardar ranking: {e}")
    
    def agregar_al_ranking(self):
        """Agrega la puntuación actual al ranking"""
        total_preguntas = len(self.resultados_detallados)
        aciertos = sum(1 for r in self.resultados_detallados if r['correcto'])
        precision = (aciertos / total_preguntas * 100) if total_preguntas > 0 else 0
        
        entrada = {
            'nombre': self.nombre_jugador,
            'experiencia': self.experiencia,
            'puntos': self.puntos,
            'precision': round(precision, 1),
            'racha_maxima': self.racha_maxima,
            'fecha': datetime.now().strftime("%Y-%m-%d %H:%M")
        }
        
        self.ranking.append(entrada)
        self.ranking.sort(key=lambda x: x['puntos'], reverse=True)
        self.ranking = self.ranking[:10]
        self.guardar_ranking()
    
    def calcular_precision(self, poligono_jugador, poligonos_correctos):
        """Calcula precisión del polígono del jugador"""
        if len(poligono_jugador) < 3:
            return 0.0
        
        mejor_coincidencia = 0.0
        for poligono_correcto in poligonos_correctos:
            coincidencia = self.calcular_superposicion(poligono_jugador, poligono_correcto)
            if coincidencia > mejor_coincidencia:
                mejor_coincidencia = coincidencia
        
        return mejor_coincidencia
    
    def calcular_superposicion(self, poli1, poli2):
        """Calcula superposición entre dos polígonos"""
        centro1 = self.obtener_centroide(poli1)
        centro2 = self.obtener_centroide(poli2)
        
        distancia = math.sqrt((centro1[0] - centro2[0]) ** 2 + (centro1[1] - centro2[1]) ** 2)
        
        area1 = self.calcular_area_poligono(poli1)
        area2 = self.calcular_area_poligono(poli2)
        area_promedio = (area1 + area2) / 2
        distancia_maxima = math.sqrt(area_promedio)
        
        if distancia_maxima > 0:
            puntuacion_distancia = max(0, 1 - (distancia / distancia_maxima))
        else:
            puntuacion_distancia = 0
        
        if max(area1, area2) > 0:
            puntuacion_area = min(area1, area2) / max(area1, area2)
        else:
            puntuacion_area = 0
        
        precision = (puntuacion_distancia * 0.6 + puntuacion_area * 0.4) * 100
        return precision
    
    def obtener_centroide(self, poligono):
        """Calcula el centroide de un polígono"""
        suma_x = sum(p['x'] if isinstance(p, dict) else p[0] for p in poligono)
        suma_y = sum(p['y'] if isinstance(p, dict) else p[1] for p in poligono)
        n = len(poligono)
        return (suma_x / n, suma_y / n)
    
    def calcular_area_poligono(self, poligono):
        """Calcula área usando fórmula de Gauss"""
        area = 0.0
        n = len(poligono)
        
        for i in range(n):
            j = (i + 1) % n
            
            if isinstance(poligono[i], dict):
                x1, y1 = poligono[i]['x'], poligono[i]['y']
                x2, y2 = poligono[j]['x'], poligono[j]['y']
            else:
                x1, y1 = poligono[i]
                x2, y2 = poligono[j]
            
            area += x1 * y2
            area -= x2 * y1
        
        return abs(area / 2.0)
    
    def iniciar_juego(self):
        """Inicia una nueva partida"""
        if self.experiencia == EXPERIENCIA_PRINCIPIANTE:
            datos_filtrados = [d for d in self.datos_juego if d['difficulty'] in [DIFICULTAD_FACIL, DIFICULTAD_MEDIA]]
        else:
            datos_filtrados = [d for d in self.datos_juego if d['difficulty'] in [DIFICULTAD_MEDIA, DIFICULTAD_DIFICIL]]
        
        random.shuffle(datos_filtrados)
        self.datos_juego = datos_filtrados
        
        print(f"\n🎲 Imágenes aleatorizadas: {len(self.datos_juego)}")
        
        self.puntos = 0
        self.vidas = 5
        self.racha = 0
        self.racha_maxima = 0
        self.pregunta_actual = 0
        self.resultados_detallados = []
        self.puntos_poligono = []
        self.respondida = False
        self.mostrar_feedback = False
        self.tiempo_inicio = pygame.time.get_ticks() / 1000
        self.estado = ESTADO_JUGANDO
    
    def enviar_respuesta(self):
        """Procesa la respuesta del jugador"""
        pregunta = self.datos_juego[self.pregunta_actual]
        poligonos_correctos = pregunta['polygons']
        es_caso_negativo = pregunta.get('es_negativo', False)
        
        puntos_ganados = 0
        es_correcto = False
        mensaje = ""
        precision = 0.0
        
        if es_caso_negativo:
            if len(self.puntos_poligono) < 3:
                es_correcto = True
                precision = 100.0
                puntos_ganados = 100
                mensaje = "¡Excelente! Identificaste correctamente que NO hay caries "
                
                self.racha += 1
                if self.racha > self.racha_maxima:
                    self.racha_maxima = self.racha
                
                if self.racha >= 3:
                    bonus_racha = self.racha * 5
                    puntos_ganados += bonus_racha
                    mensaje += f"Racha x{self.racha} (+{bonus_racha})"
                
                self.puntos += puntos_ganados
            else:
                mensaje = "Incorrecto. Esta imagen NO tiene caries (falso positivo)"
                self.vidas -= 1
                self.racha = 0
        else:
            if len(self.puntos_poligono) < 3:
                mensaje = "Incorrecto. Hay caries pero no las marcaste"
                self.vidas -= 1
                self.racha = 0
            else:
                precision = self.calcular_precision(self.puntos_poligono, poligonos_correctos)
                
                if precision >= 80:
                    es_correcto = True
                    puntos_ganados = int(precision)
                    
                    tiempo_transcurrido = (pygame.time.get_ticks() / 1000) - self.tiempo_inicio
                    tiempo_por_pregunta = tiempo_transcurrido / (self.pregunta_actual + 1)
                    
                    if tiempo_por_pregunta < 30:
                        bonus_velocidad = int((30 - tiempo_por_pregunta) / 2)
                        puntos_ganados += bonus_velocidad
                        mensaje = f"¡Excelente! +{puntos_ganados} puntos ({bonus_velocidad} bonus velocidad)"
                    else:
                        mensaje = f"¡Correcto! +{puntos_ganados} puntos"
                    
                    self.racha += 1
                    if self.racha > self.racha_maxima:
                        self.racha_maxima = self.racha
                    
                    if self.racha >= 3:
                        bonus_racha = self.racha * 5
                        puntos_ganados += bonus_racha
                        mensaje += f"Racha x{self.racha} (+{bonus_racha})"
                    
                    self.puntos += puntos_ganados
                else:
                    mensaje = f"Incorrecto. Precisión: {precision:.1f}%"
                    self.vidas -= 1
                    self.racha = 0
        
        self.precision_actual = precision
        self.es_correcto = es_correcto
        self.mensaje_feedback = mensaje
        self.respondida = True
        self.mostrar_feedback = True
        
        self.resultados_detallados.append({'pregunta': self.pregunta_actual + 1, 'correcto': es_correcto, 'precision': round(precision, 1), 'puntos': puntos_ganados})
        
        pygame.time.set_timer(pygame.USEREVENT, 3000)
    
    def siguiente_pregunta(self):
        """Avanza a la siguiente pregunta"""
        if self.vidas <= 0:
            self.terminar_juego()
            return
        
        if self.pregunta_actual + 1 >= len(self.datos_juego):
            self.terminar_juego()
            return
        
        self.pregunta_actual += 1
        self.puntos_poligono = []
        self.respondida = False
        self.mostrar_feedback = False
    
    def terminar_juego(self):
        """Finaliza el juego"""
        self.tiempo_actual = (pygame.time.get_ticks() / 1000) - self.tiempo_inicio
        self.agregar_al_ranking()
        self.guardar_partida_excel()
        self.estado = ESTADO_RESULTADOS
    
    def dibujar_menu(self):
        """Dibuja el menú principal"""
        # Fondo blanco limpio
        self.ventana.fill((255, 255, 255))
        
        # Alineación consistente para todo el contenido - más a la izquierda
        nombre_x = 200
        
        # Icono de diente alineado con el contenido
        diente_x = nombre_x
        diente_y = 60
        self.dibujar_diente(diente_x, diente_y, 60, COLOR_AZUL)
        
        # Título principal alineado con el contenido
        titulo1 = self.fuente_titulo.render("¿Puedes detectar caries", True, (30, 41, 59))
        titulo2 = self.fuente_titulo.render("en imágenes NIRI?", True, (30, 41, 59))
        self.ventana.blit(titulo1, (diente_x + 80, diente_y + 5))
        self.ventana.blit(titulo2, (diente_x + 80, diente_y + 50))
        
        # Subtítulo alineado a la izquierda con margen
        subtitulo = self.fuente_pequena.render("Entrena tus habilidades de diagnóstico visual", True, COLOR_GRIS)
        self.ventana.blit(subtitulo, (nombre_x, 160))
        
        # Contenedor de formulario
        y_actual = 220
        
        # Nombre
        texto_nombre = self.fuente_mediana.render("Nombre", True, (30, 41, 59))
        self.ventana.blit(texto_nombre, (nombre_x, y_actual))
        
        rect_input = pygame.Rect(nombre_x, y_actual + 40, 480, 55)
        color_input = COLOR_AZUL if self.input_activo else (226, 232, 240)
        
        # Input con fondo blanco y borde sutil
        pygame.draw.rect(self.ventana, (255, 255, 255), rect_input, border_radius=12)
        pygame.draw.rect(self.ventana, color_input, rect_input, 2, border_radius=12)
        
        texto_input = self.fuente_mediana.render(self.nombre_jugador, True, (30, 41, 59))
        self.ventana.blit(texto_input, (nombre_x + 20, y_actual + 55))
        self.rect_input = rect_input
        
        y_actual += 130
        texto_experiencia = self.fuente_mediana.render("Nivel de experiencia", True, (30, 41, 59))
        self.ventana.blit(texto_experiencia, (nombre_x, y_actual))
        
        y_actual += 45
        card_x = nombre_x
        card_width = 230
        card_height = 75
        
        rect_principiante = pygame.Rect(card_x, y_actual, card_width, card_height)
        es_seleccionado_prin = self.experiencia == EXPERIENCIA_PRINCIPIANTE
        
        # Card de principiante
        if es_seleccionado_prin:
            pygame.draw.rect(self.ventana, (239, 246, 255), rect_principiante, border_radius=12)
            pygame.draw.rect(self.ventana, COLOR_AZUL, rect_principiante, 2, border_radius=12)
        else:
            pygame.draw.rect(self.ventana, (255, 255, 255), rect_principiante, border_radius=12)
            pygame.draw.rect(self.ventana, (226, 232, 240), rect_principiante, 2, border_radius=12)
        
        # Icono de diente pequeño
        self.dibujar_diente(card_x + 15, y_actual + 20, 25, COLOR_AZUL if es_seleccionado_prin else COLOR_GRIS)
        
        texto_prin = self.fuente_mediana.render("0-5 años", True, (30, 41, 59))
        self.ventana.blit(texto_prin, (card_x + 50, y_actual + 25))
        self.rect_principiante = rect_principiante
        
        rect_avanzado = pygame.Rect(card_x + card_width + 20, y_actual, card_width, card_height)
        es_seleccionado_avan = self.experiencia == EXPERIENCIA_AVANZADO
        
        # Card de avanzado
        if es_seleccionado_avan:
            pygame.draw.rect(self.ventana, (239, 246, 255), rect_avanzado, border_radius=12)
            pygame.draw.rect(self.ventana, COLOR_AZUL, rect_avanzado, 2, border_radius=12)
        else:
            pygame.draw.rect(self.ventana, (255, 255, 255), rect_avanzado, border_radius=12)
            pygame.draw.rect(self.ventana, (226, 232, 240), rect_avanzado, 2, border_radius=12)
        
        # Icono de diente pequeño
        self.dibujar_diente(card_x + card_width + 35, y_actual + 20, 25, COLOR_AZUL if es_seleccionado_avan else COLOR_GRIS)
        
        texto_avan = self.fuente_mediana.render("5+ años", True, (30, 41, 59))
        self.ventana.blit(texto_avan, (card_x + card_width + 70, y_actual + 25))
        self.rect_avanzado = rect_avanzado
        
        y_actual += 110
        texto_instrucciones = self.fuente_mediana.render("Instrucciones", True, (30, 41, 59))
        self.ventana.blit(texto_instrucciones, (nombre_x, y_actual))
        
        # Panel de instrucciones con fondo blanco 
        instrucciones_rect = pygame.Rect(nombre_x, y_actual + 40, 480, 135)
        pygame.draw.rect(self.ventana, (255, 255, 255), instrucciones_rect, border_radius=12)
        pygame.draw.rect(self.ventana, (226, 232, 240), instrucciones_rect, 2, border_radius=12)

        instrucciones = [
            "• Dibuja polígonos alrededor de posibles caries",
            "• Gana puntos por precisión y velocidad",
            "• Mantén una racha para obtener bonificaciones"
        ]

        for i, linea in enumerate(instrucciones):
            texto = self.fuente_pequena.render(linea, True, (71, 85, 105))
            self.ventana.blit(texto, (nombre_x + 20, y_actual + 60 + i * 35))
        
        y_actual += 210
        
        puede_iniciar = (len(self.nombre_jugador) > 0 and self.experiencia is not None and len(self.datos_juego) > 0)
        
        rect_boton = pygame.Rect(nombre_x, y_actual, 480, 60)
        mouse_pos = pygame.mouse.get_pos()
        es_hover = rect_boton.collidepoint(mouse_pos) if puede_iniciar else False
        
        # Botón azul como en el mockup
        if puede_iniciar:
            color_boton = (37, 99, 235) if es_hover else COLOR_AZUL
            pygame.draw.rect(self.ventana, color_boton, rect_boton, border_radius=12)
        else:
            pygame.draw.rect(self.ventana, (203, 213, 225), rect_boton, border_radius=12)
        
        texto_boton = self.fuente_grande.render("Iniciar Juego", True, COLOR_BLANCO)
        rect_texto = texto_boton.get_rect(center=rect_boton.center)
        self.ventana.blit(texto_boton, rect_texto)
        self.rect_iniciar = rect_boton if puede_iniciar else None
        
        if len(self.ranking) > 0:
            y_ranking = 280
            x_ranking = ANCHO_VENTANA - 480
            
            # Panel de ranking con fondo blanco - más grande
            ranking_panel = pygame.Rect(x_ranking - 20, y_ranking - 20, 340, 450)
            pygame.draw.rect(self.ventana, (255, 255, 255), ranking_panel, border_radius=16)
            pygame.draw.rect(self.ventana, (226, 232, 240), ranking_panel, 2, border_radius=16)
            
            titulo_ranking = self.fuente_grande.render("TOP 5", True, (30, 41, 59))
            self.ventana.blit(titulo_ranking, (x_ranking, y_ranking))
            
            for i, entrada in enumerate(self.ranking[:5]):
                y_pos = y_ranking + 60 + i * 70
                
                nombre_texto = self.fuente_mediana.render(entrada['nombre'], True, (30, 41, 59))
                self.ventana.blit(nombre_texto, (x_ranking, y_pos))
                
                precision_texto = self.fuente_mediana.render(f"{entrada['precision']}%", True, (71, 85, 105))
                self.ventana.blit(precision_texto, (x_ranking + 230, y_pos))
                
                # Línea separadora sutil
                if i < len(self.ranking[:5]) - 1:
                    pygame.draw.line(self.ventana, (226, 232, 240), 
                                   (x_ranking, y_pos + 45), 
                                   (x_ranking + 300, y_pos + 45), 1)
    
    def dibujar_jugando(self):
        """Dibuja la pantalla de juego - versión simplificada para ahorrar espacio"""
        # Fondo blanco limpio como el menú
        self.ventana.fill((255, 255, 255))
        
        pregunta = self.datos_juego[self.pregunta_actual]
        nombre_imagen = pregunta['imageName']
        
        # Panel superior moderno con stats - fondo azul
        stats_panel = pygame.Rect(20, 20, ANCHO_VENTANA - 40, 80)
        pygame.draw.rect(self.ventana, (239, 246, 255), stats_panel, border_radius=12)
        pygame.draw.rect(self.ventana, COLOR_AZUL, stats_panel, 2, border_radius=12)
        
        # Centrar verticalmente todos los elementos
        panel_center_y = stats_panel.centery
        
        # Calcular espacio disponible y distribuir elementos uniformemente
        panel_width = stats_panel.width
        padding = 40  # Padding desde los bordes
        num_sections = 4  # Pregunta, Puntos, Vidas, Tiempo
        section_width = (panel_width - 2 * padding) / num_sections
        
        # Sección 1: Pregunta con indicador visual
        pregunta_x_base = stats_panel.left + padding
        pregunta_y = panel_center_y - 20
        pygame.draw.circle(self.ventana, COLOR_AZUL, (pregunta_x_base + 25, panel_center_y), 12)
        pregunta_mini = self.fuente_mediana.render(str(self.pregunta_actual + 1), True, COLOR_BLANCO)
        pregunta_mini_rect = pregunta_mini.get_rect(center=(pregunta_x_base + 25, panel_center_y))
        self.ventana.blit(pregunta_mini, pregunta_mini_rect)
        pregunta_texto = self.fuente_grande.render(f"/{len(self.datos_juego)}", True, (30, 41, 59))
        self.ventana.blit(pregunta_texto, (pregunta_x_base + 42, pregunta_y))
        
        # Sección 2: Puntos con estrella y etiqueta
        star_x_base = int(pregunta_x_base + section_width + 20)
        star_y_base = panel_center_y - 15
        
        # Label encima y a la izquierda de la estrella
        puntos_label = self.fuente_pequena.render("Puntos:", True, (71, 85, 105))
        self.ventana.blit(puntos_label, (star_x_base - 15, panel_center_y - 28))
        
        star_points = [
            (star_x_base, star_y_base + 5),
            (star_x_base + 5, star_y_base + 15),
            (star_x_base + 15, star_y_base + 15),
            (star_x_base + 7, star_y_base + 20),
            (star_x_base + 10, star_y_base + 30),
            (star_x_base, star_y_base + 25),
            (star_x_base - 10, star_y_base + 30),
            (star_x_base - 7, star_y_base + 20),
            (star_x_base - 15, star_y_base + 15),
            (star_x_base - 5, star_y_base + 15)
        ]
        pygame.draw.polygon(self.ventana, COLOR_AMARILLO, star_points)
        puntos_texto = self.fuente_mediana.render(f"{self.puntos}", True, (30, 41, 59))
        self.ventana.blit(puntos_texto, (star_x_base + 25, panel_center_y - 2))
        
        # Sección 3: Vidas con corazones
        vida_label_x = int(pregunta_x_base + 2 * section_width)
        vida_texto = self.fuente_pequena.render("Vidas:", True, (71, 85, 105))
        self.ventana.blit(vida_texto, (vida_label_x, panel_center_y - 23))
        
        vida_x = vida_label_x
        vida_y = panel_center_y - 2
        heart_size = 24
        for i in range(5):
            heart_x = vida_x + i * 28
            if i < self.vidas:
                # Corazón rojo lleno
                self.dibujar_corazon(heart_x, vida_y, heart_size, COLOR_ROJO, filled=True)
            else:
                # Corazón gris vacío
                self.dibujar_corazon(heart_x, vida_y, heart_size, (70, 70, 70), filled=True)
        
        # Sección 4: Tiempo con icono de reloj
        clock_x = int(pregunta_x_base + 3 * section_width)
        clock_label_texto = self.fuente_pequena.render("Tiempo:", True, (71, 85, 105))
        self.ventana.blit(clock_label_texto, (clock_x, panel_center_y - 23))
        
        clock_icon_x, clock_icon_y = clock_x + 10, panel_center_y + 8
        pygame.draw.circle(self.ventana, (30, 41, 59), (clock_icon_x, clock_icon_y), 12, 2)
        pygame.draw.line(self.ventana, (30, 41, 59), (clock_icon_x, clock_icon_y), (clock_icon_x, clock_icon_y - 6), 2)
        pygame.draw.line(self.ventana, (30, 41, 59), (clock_icon_x, clock_icon_y), (clock_icon_x + 5, clock_icon_y), 2)
        
        tiempo = int((pygame.time.get_ticks() / 1000) - self.tiempo_inicio)
        tiempo_texto = self.fuente_mediana.render(f"{tiempo//60}:{tiempo%60:02d}", True, (30, 41, 59))
        self.ventana.blit(tiempo_texto, (clock_x + 35, panel_center_y - 2))
        
        if self.racha >= 3:
            # Racha con diseño flat moderno
            racha_rect = pygame.Rect(ANCHO_VENTANA // 2 - 140, 110, 280, 60)
            self.dibujar_panel_moderno(racha_rect, COLOR_AMARILLO, radio=30, sombra=True, intensidad_sombra=50)
            
            texto_racha = self.fuente_grande.render(f"RACHA x{self.racha}!", True, COLOR_NEGRO)
            self.ventana.blit(texto_racha, texto_racha.get_rect(center=racha_rect.center))
        
        y_imagen = 140 if self.racha < 3 else 210
        
        if nombre_imagen in self.imagenes_cargadas:
            imagen = self.imagenes_cargadas[nombre_imagen]
            # Más espacio para la imagen y mejor centrado
            max_width = 850
            max_height = ALTO_VENTANA - y_imagen - 30
            escala = min(max_width / imagen.get_width(), max_height / imagen.get_height(), 1.0)
            imagen_escalada = pygame.transform.scale(imagen, (int(imagen.get_width() * escala), int(imagen.get_height() * escala)))
            
            # Centrar la imagen en el espacio disponible
            espacio_disponible = ANCHO_VENTANA - 480 - 150  # Entre contenido izquierdo y panel derecho
            x_centrado = 150 + (espacio_disponible - imagen_escalada.get_width()) // 2
            rect_imagen = imagen_escalada.get_rect(topleft=(x_centrado, y_imagen))
            self.ventana.blit(imagen_escalada, rect_imagen)
            
            if len(self.puntos_poligono) > 0:
                puntos_pantalla = [(rect_imagen.left + p[0] * escala, rect_imagen.top + p[1] * escala) for p in self.puntos_poligono]
                if len(puntos_pantalla) > 1:
                    pygame.draw.lines(self.ventana, COLOR_AZUL, False, puntos_pantalla, 3)
                for i, punto in enumerate(puntos_pantalla):
                    pygame.draw.circle(self.ventana, COLOR_VERDE if i == 0 else COLOR_AZUL, punto, 6)
            
            if self.mostrar_feedback and not pregunta.get('es_negativo', False):
                for poligono_correcto in pregunta['polygons']:
                    puntos_correctos = [(rect_imagen.left + p['x'] * escala, rect_imagen.top + p['y'] * escala) for p in poligono_correcto]
                    if len(puntos_correctos) > 2:
                        pygame.draw.polygon(self.ventana, COLOR_VERDE, puntos_correctos, 3)
            
            self.rect_imagen = rect_imagen
            self.escala_imagen = escala
        
              # *** PANEL LATERAL DERECHO ***
        x_panel = ANCHO_VENTANA - 480
        y_panel = y_imagen
        ancho_panel = 360
        
        # Panel lateral moderno - fondo azul
        panel_rect = pygame.Rect(x_panel, y_panel, ancho_panel, ALTO_VENTANA - y_panel - 30)
        pygame.draw.rect(self.ventana, (239, 246, 255), panel_rect, border_radius=16)
        pygame.draw.rect(self.ventana, COLOR_AZUL, panel_rect, 2, border_radius=16)
        
        # Título del panel
        titulo_panel = self.fuente_mediana.render("Tu Progreso", True, (30, 41, 59))
        self.ventana.blit(titulo_panel, (x_panel + 20, y_panel + 20))
        
        y_info = y_panel + 70
        
        # Dificultad de la pregunta
        dificultad = pregunta['difficulty']
        color_dificultad = (
            COLOR_VERDE if dificultad == DIFICULTAD_FACIL 
            else COLOR_AMARILLO if dificultad == DIFICULTAD_MEDIA 
            else COLOR_ROJO
        )
        
        dificultad_rect = pygame.Rect(x_panel + 20, y_info, 320, 40)
        pygame.draw.rect(self.ventana, color_dificultad, dificultad_rect, border_radius=12)
        
        texto_dificultad = self.fuente_pequena.render(
            f"Dificultad: {dificultad.upper()}", 
            True, 
            COLOR_BLANCO
        )
        self.ventana.blit(texto_dificultad, (x_panel + 30, y_info + 10))
        
        # Racha actual
        y_info += 60
        texto_racha_label = self.fuente_pequena.render("Racha actual:", True, (71, 85, 105))
        self.ventana.blit(texto_racha_label, (x_panel + 20, y_info))
        texto_racha_valor = self.fuente_mediana.render(str(self.racha), True, COLOR_AMARILLO)
        self.ventana.blit(texto_racha_valor, (x_panel + 270, y_info - 3))
        
        # Mejor racha
        y_info += 50
        texto_max_racha_label = self.fuente_pequena.render("Mejor racha:", True, (71, 85, 105))
        self.ventana.blit(texto_max_racha_label, (x_panel + 20, y_info))
        texto_max_racha_valor = self.fuente_mediana.render(
            str(self.racha_maxima), 
            True, 
            COLOR_VERDE
        )
        self.ventana.blit(texto_max_racha_valor, (x_panel + 270, y_info - 3))
        
        # Puntos marcados
        y_info += 50
        texto_puntos_label = self.fuente_pequena.render("Puntos marcados:", True, (71, 85, 105))
        self.ventana.blit(texto_puntos_label, (x_panel + 20, y_info))
        texto_puntos_valor = self.fuente_mediana.render(
            str(len(self.puntos_poligono)), 
            True, 
            COLOR_AZUL
        )
        self.ventana.blit(texto_puntos_valor, (x_panel + 270, y_info - 3))
        
        # Instrucciones simplificadas
        y_info += 70
        instrucciones = [
            "Controles:",
            "• Click: añadir puntos",
            "• ESPACIO: deshacer",
            "• ENTER: limpiar todo"
        ]
        
        for i, linea in enumerate(instrucciones):
            fuente = self.fuente_pequena
            color = (30, 41, 59) if i == 0 else (71, 85, 105)
            texto = fuente.render(linea, True, color)
            self.ventana.blit(texto, (x_panel + 20, y_info + i * 30))
        
        # Botones de envío - diseño flat moderno
        y_botones = ALTO_VENTANA - 250
        puede_enviar = len(self.puntos_poligono) >= 3 and not self.respondida
        puede_enviar_vacio = len(self.puntos_poligono) == 0 and not self.respondida
        mouse_pos = pygame.mouse.get_pos()
        
        # Botón ENVIAR (para cuando hay polígono)
        rect_enviar = pygame.Rect(x_panel + 20, y_botones + 110, 320, 60)
        es_hover_enviar = rect_enviar.collidepoint(mouse_pos) if puede_enviar else False
        
        if puede_enviar:
            color_enviar = COLOR_VERDE_HOVER if es_hover_enviar else COLOR_VERDE
            self.dibujar_panel_moderno(rect_enviar, color_enviar, radio=16, sombra=True, intensidad_sombra=50)
            if es_hover_enviar:
                # Glow sutil
                glow_surf = pygame.Surface((rect_enviar.width + 4, rect_enviar.height + 4), pygame.SRCALPHA)
                pygame.draw.rect(glow_surf, (*COLOR_VERDE, 60), glow_surf.get_rect(), border_radius=18)
                self.ventana.blit(glow_surf, (rect_enviar.x - 2, rect_enviar.y - 2))
        else:
            pygame.draw.rect(self.ventana, COLOR_GRIS_OSCURO, rect_enviar, 2, border_radius=16)
        
        color_texto_enviar = COLOR_BLANCO if puede_enviar else COLOR_GRIS
        texto_enviar = self.fuente_mediana.render("Enviar" if puede_enviar else "Esperando...", True, color_texto_enviar)
        self.ventana.blit(texto_enviar, texto_enviar.get_rect(center=rect_enviar.center))
        self.rect_enviar = rect_enviar if puede_enviar else None
        
        # Botón SIN CARIES (siempre visible cuando no se ha respondido)
        rect_sin_caries = pygame.Rect(x_panel + 20, y_botones + 40, 320, 55)
        es_hover_sin_caries = rect_sin_caries.collidepoint(mouse_pos) if not self.respondida else False
        
        if not self.respondida:
            color_sin_caries = COLOR_AZUL_CLARO if es_hover_sin_caries else COLOR_AZUL
            pygame.draw.rect(self.ventana, color_sin_caries, rect_sin_caries, border_radius=12)
            if es_hover_sin_caries:
                glow_surf = pygame.Surface((rect_sin_caries.width + 4, rect_sin_caries.height + 4), pygame.SRCALPHA)
                pygame.draw.rect(glow_surf, (*COLOR_AZUL, 60), glow_surf.get_rect(), border_radius=14)
                self.ventana.blit(glow_surf, (rect_sin_caries.x - 2, rect_sin_caries.y - 2))
        else:
            pygame.draw.rect(self.ventana, COLOR_GRIS_OSCURO, rect_sin_caries, 2, border_radius=12)
        
        color_texto_sin_caries = COLOR_BLANCO if not self.respondida else COLOR_GRIS
        texto_sin_caries = self.fuente_mediana.render("Sin Caries", True, color_texto_sin_caries)
        self.ventana.blit(texto_sin_caries, texto_sin_caries.get_rect(center=rect_sin_caries.center))
        self.rect_sin_caries = rect_sin_caries if not self.respondida else None
        
        if self.mostrar_feedback:
            # Animación de feedback - flat design
            self.feedback_timer += 1
            self.feedback_alpha = min(255, self.feedback_timer * 25)
            
            feedback_rect = pygame.Rect(ANCHO_VENTANA // 2 - 400, ALTO_VENTANA // 2 - 80, 800, 160)
            
            superficie_feedback = pygame.Surface((feedback_rect.width, feedback_rect.height), pygame.SRCALPHA)
            color_base = COLOR_VERDE if self.es_correcto else COLOR_ROJO
            color_con_alpha = (*color_base, min(230, self.feedback_alpha))
            pygame.draw.rect(superficie_feedback, color_con_alpha, superficie_feedback.get_rect(), border_radius=24)
            
            self.ventana.blit(superficie_feedback, (feedback_rect.x, feedback_rect.y))
            
            if self.feedback_alpha > 50:
                texto_mensaje = self.fuente_grande.render(self.mensaje_feedback, True, COLOR_BLANCO)
                self.ventana.blit(texto_mensaje, texto_mensaje.get_rect(center=(ANCHO_VENTANA // 2, ALTO_VENTANA // 2 - 25)))
                texto_precision = self.fuente_mediana.render(f"Precisión: {self.precision_actual:.1f}%", True, COLOR_BLANCO)
                self.ventana.blit(texto_precision, texto_precision.get_rect(center=(ANCHO_VENTANA // 2, ALTO_VENTANA // 2 + 20)))
        else:
            self.feedback_timer = 0
            self.feedback_alpha = 0
    
    def dibujar_resultados(self):
        """Dibuja la pantalla de resultados"""
        # Fondo blanco limpio como el menú
        self.ventana.fill((255, 255, 255))
        
        total_preguntas = len(self.resultados_detallados)
        aciertos = sum(1 for r in self.resultados_detallados if r['correcto'])
        precision = (aciertos / total_preguntas * 100) if total_preguntas > 0 else 0
        
        if precision >= 90: nivel, icono, color_nivel = "EXPERTO MAESTRO", " ", COLOR_AMARILLO
        elif precision >= 80: nivel, icono, color_nivel = "EXPERTO", " ", COLOR_VERDE
        elif precision >= 70: nivel, icono, color_nivel = "AVANZADO", " ", COLOR_AZUL
        elif precision >= 60: nivel, icono, color_nivel = "INTERMEDIO", " ", (168, 85, 247)
        else: nivel, icono, color_nivel = "PRINCIPIANTE", " ", COLOR_GRIS
        
        self.ventana.blit(self.fuente_titulo.render(icono, True, (30, 41, 59)), self.fuente_titulo.render(icono, True, (30, 41, 59)).get_rect(center=(ANCHO_VENTANA // 2, 80)))
        self.ventana.blit(self.fuente_titulo.render("¡Juego Completado!", True, (30, 41, 59)), self.fuente_titulo.render("¡Juego Completado!", True, (30, 41, 59)).get_rect(center=(ANCHO_VENTANA // 2, 170)))
        self.ventana.blit(self.fuente_grande.render(nivel, True, color_nivel), self.fuente_grande.render(nivel, True, color_nivel).get_rect(center=(ANCHO_VENTANA // 2, 230)))
        self.ventana.blit(self.fuente_pequena.render("Datos guardados en Excel", True, COLOR_VERDE), self.fuente_pequena.render("Datos guardados en Excel", True, COLOR_VERDE).get_rect(center=(ANCHO_VENTANA // 2, 270)))
        
        rect_volver = pygame.Rect(ANCHO_VENTANA // 2 - 180, ALTO_VENTANA - 120, 360, 70)
        mouse_pos = pygame.mouse.get_pos()
        es_hover_volver = rect_volver.collidepoint(mouse_pos)
        
        color_volver = COLOR_AZUL_CLARO if es_hover_volver else COLOR_AZUL
        self.dibujar_panel_moderno(rect_volver, color_volver, radio=20, sombra=True, intensidad_sombra=60)
        
        if es_hover_volver:
            glow_surf = pygame.Surface((rect_volver.width + 6, rect_volver.height + 6), pygame.SRCALPHA)
            pygame.draw.rect(glow_surf, (*COLOR_AZUL, 70), glow_surf.get_rect(), border_radius=22)
            self.ventana.blit(glow_surf, (rect_volver.x - 3, rect_volver.y - 3))
        
        texto_volver = self.fuente_grande.render("Volver al menú", True, COLOR_BLANCO)
        self.ventana.blit(texto_volver, texto_volver.get_rect(center=rect_volver.center))
        self.rect_volver = rect_volver
    
    def manejar_eventos(self):
        """Procesa eventos del usuario"""
        for evento in pygame.event.get():
            if evento.type == pygame.QUIT:
                return False
            
            if evento.type == pygame.USEREVENT and self.estado == ESTADO_JUGANDO and self.respondida:
                self.siguiente_pregunta()
            
            if self.estado == ESTADO_MENU:
                if evento.type == pygame.MOUSEBUTTONDOWN:
                    pos = evento.pos
                    if hasattr(self, 'rect_input') and self.rect_input.collidepoint(pos): self.input_activo = True
                    else: self.input_activo = False
                    if hasattr(self, 'rect_principiante') and self.rect_principiante.collidepoint(pos): self.experiencia = EXPERIENCIA_PRINCIPIANTE
                    if hasattr(self, 'rect_avanzado') and self.rect_avanzado.collidepoint(pos): self.experiencia = EXPERIENCIA_AVANZADO
                    if hasattr(self, 'rect_iniciar') and self.rect_iniciar and self.rect_iniciar.collidepoint(pos): self.iniciar_juego()
                
                if evento.type == pygame.KEYDOWN and self.input_activo:
                    if evento.key == pygame.K_BACKSPACE: self.nombre_jugador = self.nombre_jugador[:-1]
                    elif evento.key == pygame.K_RETURN: self.input_activo = False
                    elif len(self.nombre_jugador) < 20: self.nombre_jugador += evento.unicode
            
            elif self.estado == ESTADO_JUGANDO:
                if evento.type == pygame.MOUSEBUTTONDOWN:
                    pos = evento.pos
                    if not self.respondida and hasattr(self, 'rect_imagen') and self.rect_imagen.collidepoint(pos):
                        x = (pos[0] - self.rect_imagen.left) / self.escala_imagen
                        y = (pos[1] - self.rect_imagen.top) / self.escala_imagen
                        self.puntos_poligono.append((x, y))
                    if hasattr(self, 'rect_enviar') and self.rect_enviar and self.rect_enviar.collidepoint(pos): self.enviar_respuesta()
                    if hasattr(self, 'rect_sin_caries') and self.rect_sin_caries and self.rect_sin_caries.collidepoint(pos): self.enviar_respuesta()
                
                if evento.type == pygame.KEYDOWN:
                    if evento.key == pygame.K_SPACE and not self.respondida and len(self.puntos_poligono) > 0: self.puntos_poligono.pop()
                    if evento.key == pygame.K_RETURN and not self.respondida: self.puntos_poligono = []
                    if evento.key == pygame.K_e and not self.respondida: self.enviar_respuesta()
            
            elif self.estado == ESTADO_RESULTADOS:
                if evento.type == pygame.MOUSEBUTTONDOWN and hasattr(self, 'rect_volver') and self.rect_volver.collidepoint(evento.pos):
                    self.estado = ESTADO_MENU
                    self.nombre_jugador = ""
                    self.experiencia = None
        
        return True
    
    def ejecutar(self):
        """Bucle principal del juego"""
        ejecutando = True
        while ejecutando:
            ejecutando = self.manejar_eventos()
            if self.estado == ESTADO_MENU: self.dibujar_menu()
            elif self.estado == ESTADO_JUGANDO: self.dibujar_jugando()
            elif self.estado == ESTADO_RESULTADOS: self.dibujar_resultados()
            pygame.display.flip()
            self.reloj.tick(60)
        pygame.quit()
        sys.exit()

if __name__ == "__main__":
    print("=" * 70)
    print("🦷 JUEGO DE DETECCIÓN DE CARIES EN IMÁGENES NIRI")
    print("=" * 70)
    print("\n✨ CARACTERÍSTICAS:")
    print("   ✅ Imágenes en orden aleatorio por partida")
    print("   ✅ Música de fondo (soundtrak_caries.mp3)")
    print("   ✅ Imagen de fondo (fondo_pantalla.jpg)")
    print("   ✅ Guardado automático en Excel (datos_juego_caries.xlsx)")
    print("   ✅ Ranking persistente en JSON")
    print("\n📋 ARCHIVOS NECESARIOS:")
    print("   • etiquetas_caries.json (imágenes etiquetadas)")
    print("   • imagenes/ (carpeta con imágenes)")
    print("   • fondo_pantalla.jpg (opcional)")
    print("   • soundtrak_caries.mp3 (opcional)")
    print("\n💾 INSTALACIÓN:")
    print("   pip install pygame openpyxl")
    print("\n" + "=" * 70)
    
    print("\n🔍 VERIFICANDO ARCHIVOS...")
    archivos_criticos = ['etiquetas_caries.json']
    falta_critico = False
    
    for archivo in archivos_criticos:
        if os.path.exists(archivo):
            print(f"   ✅ {archivo}")
        else:
            print(f"   ❌ {archivo} (NECESARIO)")
            falta_critico = True
    
    for archivo in ['fondo_pantalla.jpg', 'soundtrak_caries.mp3']:
        print(f"   {'✅' if os.path.exists(archivo) else '⚠️ '} {archivo} ({'encontrado' if os.path.exists(archivo) else 'opcional'})")
    
    if os.path.exists('imagenes'):
        num = len([f for f in os.listdir('imagenes') if f.endswith(('.jpg', '.jpeg', '.png'))])
        print(f"   ✅ Carpeta imagenes/ ({num} imágenes)")
    else:
        print(f"   ❌ Carpeta imagenes/ (NECESARIA)")
        falta_critico = True
    
    if falta_critico:
        print("\n⚠️  El juego se iniciará con imágenes simuladas de respaldo")
    else:
        print("\n✅ TODOS LOS ARCHIVOS ENCONTRADOS")
    
    input("\nPresiona ENTER para iniciar el juego...")
    print("\n🚀 Iniciando juego...\n")
    
    try:
        juego = JuegoDeteccionCaries()
        juego.ejecutar()
    except Exception as e:
        print(f"\n❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        input("\nPresiona ENTER para salir...")